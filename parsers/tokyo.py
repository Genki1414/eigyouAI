"""
parsers/tokyo.py — 東京都「建設業許可業者名簿」Excelパーサー
出典: 東京都都市整備局（建設業情報管理センター登録情報、月1回更新でExcel公開）
      https://www.toshiseibi.metro.tokyo.lg.jp/kenchiku_kaihatsu/kenchiku_shidou/gyosya_shido/kensetsu/tokyou_kensetsu_meibo

実ファイル(meibo_kyokajunR0806.xlsx, 2026-08受領)で確認した構造:
  1行目: 資本金・許可業種コードの注記（データではない）
  2行目: 一般項目のヘッダ（許可番号/商号名称/電話番号/郵便番号/県名市区町村名/
         所在地/資本金額/許可年月日/許可業種 など）。「許可業種」は29列に
         またがる見出しで、この行では1回しか出てこない
  3行目: 29業種それぞれの1文字略号（土/建/大/左/と/石/屋/電/管/タ/鋼/筋/ほ/し/
         板/ガ/塗/防/内/機/絶/通/園/井/具/水/消/清/解）
  4行目以降: データ。業種列の値は 0=なし/1=一般建設業/2=特定建設業（未許可は空欄）
  末尾: 全列空欄の行が続く（シート範囲の余白）

このファイルには許可行政庁(知事/大臣)の列が無い。東京都が自庁で公表する
名簿である以上、大臣許可業者（本店が複数都道府県にまたがる業者）は国交省の
地方整備局が所管し、この名簿には含まれない前提で扱う（知事許可のみ収録）。
実データでこの前提が崩れていないか、本番投入前に必ず確認すること。

許可番号は素の連番(例: 5, 8, 9...)で都道府県内でのみ一意なため、他県を
追加した際の衝突を避けて "<都道府県名>-<連番>" の形でDBのlicense_noに格納する。

4万行超をメモリに全展開しないよう、ヘッダ検出後はストリーミングで読む。

⚠ founded_year(会社の設立年)について: このファイルには設立年月日の列が無く、
  取れるのは許可年月日(=直近の許可・更新日)のみ。建設業許可は5年ごとの更新制
  のため、実データでは founded_year が2021〜2026年に偏り、実際の社歴を表さない
  （db.pyのコメント通りあくまで許可年月日からの「推定」）。scoring.pyの成長シグナル
  (founded_year>=2010で+4点)とlearn.pyのage_years特徴量は、この実データでは
  ほぼ全社が条件を満たしてしまい、意図した判別力を持たない。設立年の実データが
  別途無い限りは、scoring.py/learn.py側の対応要否を判断してもらう必要がある
  （ingest層の実装だけでは直せない）。
"""
import re

import config as C
from . import common

PREF_NAME = "東京都"

# 2行目(一般項目ヘッダ)の候補語
_HEADER_CANDIDATES = {
    "license_no": ["許可番号", "登録番号"],
    "name": ["商号名称", "商号又は名称", "商号", "名称"],
    "phone": ["電話番号", "電話"],
    "fax": ["FAX番号", "ＦＡＸ番号"],
    "pref_city": ["県名市区町村名", "都道府県市区町村"],
    "address": ["所在地"],
    "capital": ["資本金額", "資本金の額", "資本金"],
    "license_date": ["許可年月日", "許可(更新)年月日"],
    "license_type": ["許可行政庁", "許可区分", "許可の種類"],  # 実ファイルには無いが将来のため残す
    "trade_single": ["許可業種", "業種"],
}


def _norm_cells(row):
    return ["" if c is None else str(c).strip() for c in row]


def _find_trade_abbrev_row(rows, start_i):
    """rows[start_i:] から、29業種の1文字略号が並ぶ行を探し、
    (見つかった行の絶対インデックス, {category: 列番号}) を返す。
    無ければ (None, {})。"""
    for i in range(start_i, len(rows)):
        cells = _norm_cells(rows[i])
        trade_cols = {}
        for j, cell in enumerate(cells):
            cat = common.category_from_abbrev(cell)
            if cat:
                trade_cols[cat] = j
        if trade_cols:
            return i, trade_cols
    return None, {}


def parse(path):
    """東京都名簿Excelを読み、companies共通スキーマの行(dict)を対象業種の会社分だけyieldする。"""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    row_iter = ws.iter_rows(values_only=True)
    lookahead = []
    for _ in range(10):
        try:
            lookahead.append(next(row_iter))
        except StopIteration:
            break
    if not lookahead:
        print(f"警告: {path} は空です")
        return

    header_i, cols = common.find_header_row(lookahead, _HEADER_CANDIDATES)
    if header_i is None or "name" not in cols or "license_no" not in cols:
        raise ValueError(
            f"{path}: ヘッダ行(許可番号/商号名称 等)を検出できませんでした。"
            f"先頭数行の実際の内容: {lookahead[:3]}"
        )

    # 業種列の特定: まず29業種の1文字略号が並ぶ行(ヘッダの次行が典型)を探す(横持ち)。
    # 見つからなければ「許可業種」等の単一テキスト列(縦持ち)にフォールバックする。
    abbrev_i, trade_cols = _find_trade_abbrev_row(lookahead, header_i + 1)
    wide_format = bool(trade_cols)
    if not wide_format and "trade_single" not in cols:
        raise ValueError(
            f"{path}: 業種を表す列を検出できませんでした"
            f"（29業種の1文字略号の行も、許可業種/業種のテキスト列も見つからない）。"
            f"ヘッダ検出結果: {cols}"
        )
    if "license_type" not in cols:
        print(f"警告: {path} に許可行政庁/許可区分の列が無いため、大臣許可の除外ができません。"
              "東京都が自庁公表する名簿は知事許可のみを収録している前提で、"
              "全行を知事許可として扱います。実データでこの前提を確認してください。")

    # データ開始行(1-indexed)を確定。lookahead分は使い切っているので、
    # 以降は min_row を指定して未読み分だけストリーミングする。
    last_used_i = abbrev_i if wide_format else header_i
    data_start_row = last_used_i + 2  # 0-indexed→1-indexed(+1) して次の行から(+1)

    def cell(row, concept):
        j = cols.get(concept)
        return row[j] if j is not None and j < len(row) else None

    n_in = n_target = 0
    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        if row is None or all(c is None for c in row):
            continue
        n_in += 1

        license_type_raw = str(cell(row, "license_type") or "")
        if "大臣" in license_type_raw:
            continue  # 大臣許可は当面スコープ外（列が無い実ファイルでは発生しない）

        if wide_format:
            trades = ",".join(sorted(
                cat for cat, j in trade_cols.items()
                if j < len(row) and common.is_licensed_flag(row[j])))
        else:
            blob = str(cell(row, "trade_single") or "")
            trades = common.normalize_trades(re.split(r"[、,・/\n]+", blob))

        if not trades:
            continue
        n_target += 1

        pref_city_raw = cell(row, "pref_city")
        pref, city = common.split_pref_city(pref_city_raw) if pref_city_raw else (None, None)

        raw_license_no = cell(row, "license_no")
        yield {
            "license_no": f"{PREF_NAME}-{raw_license_no}" if raw_license_no is not None else None,
            "name": cell(row, "name"),
            "pref": pref or PREF_NAME,
            "city": city,
            "address": cell(row, "address"),
            "phone": cell(row, "phone"),
            "fax": cell(row, "fax"),
            "license_type": "知事",  # 大臣許可は上で除外済み（列がある場合のみ）
            "trades": trades,
            "capital": common.parse_capital(cell(row, "capital")),
            "founded_year": common.parse_year(cell(row, "license_date")),
        }

    print(f"  {PREF_NAME}: 読込{n_in}行 / 対象業種{n_target}件"
          f" ({'横持ち(略号)' if wide_format else '縦持ち'}形式で解釈)")
    if n_in > 0 and n_target == 0:
        print(f"警告: {path} から対象業種({'/'.join(C.TARGET_TRADES.keys())})の会社が1件も取れませんでした。"
              f"業種列の形式・表記を確認してください。ヘッダ検出結果: {cols}"
              f" / 横持ち候補列: {trade_cols}")
